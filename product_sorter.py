import pandas as pd 
from constant import constant
import openpyxl

from datetime import datetime 

##-- Global Variables --- Input Excel files
customer_file = constant.customer_file
order_statement = constant.order_statement
order_sheetname = constant.order_statement_sheetname

##-- Auxiliary Functions
def get_date_sheetname() -> str:
    '''
    This function determines the sheet name to look up for order statement creation
    
    Input:
        None
        
    Output:
        final_date_sheetname: str
          
    '''
    ## Today as a default value
    today = datetime.now().strftime("%d%m%y")
    
    date_to_generate = input("กรุณาใส่วันที่ที่ต้องการสร้างไฟล์ในรูปแบบ วันเดือนปี ตัวอย่างเช่น 181124 : ")
    
    date_sheetname = date_to_generate if date_to_generate != '' else today
    
    ## Monday : 0, Sunday : 6
    thai_days =  {'0': "จ.", 
                 '1': "อ.", 
                 '2': "พ.", 
                 '3': "พฤ.", 
                 '4': "ศ.", 
                 '5': "ส."
                }
    
    weekday = datetime.strptime(date_sheetname, "%d%m%y").weekday()
    
    # Concat strings to make a sheet name
    date_sheetname_final = thai_days[str(weekday)] +  date_sheetname 
    
    return date_sheetname_final

def get_customer_list(df) -> list:
    '''
    This function determines the number of output Exceil files to generate, and the list of all customers for a certain date
    
    Input:
        df_cols : dataframe
        
    Output:
        customer_list: list
          
    '''
    last_index = df.columns.get_loc('คงเหลือ')
    
    customer_list = []
    
    for column in df_cust.columns[last_index+1:]:

        if type(column) != float:
            customer_list.append(column)
                
        else:
            break

    print(f"There are {len(customer_list)} customers : {customer_list}")

    return customer_list

def extract_and_paste_product_code(df_cust, destination_sheet):
    
    row_num_end = df_cust[df_cust['รหัส'].eq('รวมปัง 5 บาท')].index  ## row number end at 32

    df_cust_code = df_cust[1:row_num_end[0]]
    
    for index, row in df_cust_code.iterrows():
        destination_sheet.cell(row = (4+index), column=2).value = row['รหัส']

def extract_and_paste_product_name(df_cust, destination_sheet):
    
    row_num_end = df_cust[df_cust['รหัส'].eq('รวมปัง 5 บาท')].index  ## row number end at 32

    df_cust_code = df_cust[1:row_num_end[0]]
    
    for index, row in df_cust_code.iterrows():
        destination_sheet.cell(row = (4+index), column=3).value = row['รายการสินค้า']


def extract_and_paste_order_quantity(df_cust, customer, destination_sheet):
    
    row_num_end = df_cust[df_cust['รหัส'].eq('รวมปัง 5 บาท')].index  ## row number end at 32

    df_cust_code = df_cust[1:row_num_end[0]]
    
    for index, row in df_cust_code.iterrows():
        destination_sheet.cell(row = (4+index), column=4).value = row[customer]



##-- Entry Point

sheetname = get_date_sheetname()

# Extract the input file 
df_cust = pd.read_excel(customer_file, sheet_name=sheetname, usecols='A:U', skiprows=2)

# Get a list of customers along with number of files
customer_list = get_customer_list(df_cust)

# Create an output Excel file for each customer 
for customer in customer_list:
    
    workbook = openpyxl.load_workbook(order_statement)
    # select the sheet to modify
    sheet = workbook[order_sheetname]

    date_sheetname = f"{sheetname[2:4]} / {sheetname[4:6]} / {sheetname[6:]}"

    # Include a date to the sheet
    sheet.cell(row=1, column=1).value = 'เอกสารจัดสินค้า เชิญชิม-บงกช (บจก.เชิญชิม ฟู้ด 1976)   วันที่..................' + date_sheetname + '..................'

    # Fill in a customer name to the form
    sheet.cell(row=2, column=1).value  = 'CODE  : C................. // Customer ...........' + customer + '...........'

    # Populate the product code
    extract_and_paste_product_code(df_cust, sheet)
    
    # Populate the product name
    extract_and_paste_product_name(df_cust, sheet)
    
    # Populate order quantity for each customer
    extract_and_paste_order_quantity(df_cust, customer, sheet)


    # save the changes
    excel_name = sheetname + customer + '.xlsx'
    workbook.save(excel_name)

    print(f"Output File [ {excel_name} ] has been created")
